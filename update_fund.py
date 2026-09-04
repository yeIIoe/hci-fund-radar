#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Atualiza o HCI FUND Radar sem bibliotecas externas.

O radar e uma camada de direcao para decisao manual. Ele nao envia ordens e
nao cria gatilhos de entrada. O FUND preserva a formula V0.1 congelada:

    spread = yield_2y(base) - yield_2y(quote)
    raw    = spread - spread[20]
    z      = (raw - media_252_ex_ante) / desvio_252_ex_ante
    FUND   = 100 * tanh(z / 2)

As medias e desvios de t usam apenas informacao ate t-1.
"""
from __future__ import annotations

import argparse
import bisect
import csv
import io
import json
import math
import re
import statistics
import subprocess
import sys
import urllib.request
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from xls_biff import read_sheet as read_biff_sheet
from next_day_model import build_next_day_observations
from pre_fund_model import build_pre_fund_radar


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
RAW_DIR = DATA_DIR / "raw"
SNAPSHOT_PATH = DATA_DIR / "fund_snapshot.json"
BACKTEST_PATH = DATA_DIR / "backtest_default.json"
CALENDAR_DIR = DATA_DIR / "calendar"
CALENDAR_INDEX_PATH = CALENDAR_DIR / "index.json"
# Caminhos dos yields "core": na maquina do Eduardo vivem em C:\Trading (fonte
# canonica local); NA NUVEM (GitHub Actions) nao existe C:\ — cai no espelho
# versionado em data/core/, que o proprio pipeline atualiza e commita.
# Override explicito por env se precisar: HCI_CORE_YIELDS / HCI_CORE_RAW.
import os as _os
_local_y = Path(r"C:\Trading\hci-ea\out\swing_fx\yields_2y.csv")
_local_r = Path(r"C:\Trading\hci-ea\out\swing_fx\raw")
CORE_YIELDS = Path(_os.environ.get("HCI_CORE_YIELDS",
                   str(_local_y if _local_y.exists() else ROOT / "data" / "core" / "yields_2y.csv")))
CORE_RAW_DIR = Path(_os.environ.get("HCI_CORE_RAW",
                    str(_local_r if _local_r.exists() else ROOT / "data" / "core" / "raw")))
CORE_RAW_DIR.mkdir(parents=True, exist_ok=True)
UST_DIR = RAW_DIR / "ust"

LOOKBACK = 20
NORM_WINDOW = 252
MIN_HISTORY = 126
FFILL_LIMIT = 5
DEFAULT_BACKTEST_START = date(2018, 1, 1)
DEFAULT_COST_PIPS = 1.0
CALENDAR_START = date(2002, 1, 1)

CURRENCIES = ("EUR", "GBP", "AUD", "NZD", "USD", "CAD", "CHF", "JPY")
PAIR_ORDER = (
    "EURGBP", "EURAUD", "EURNZD", "EURUSD", "EURCAD", "EURCHF", "EURJPY",
    "GBPAUD", "GBPNZD", "GBPUSD", "GBPCAD", "GBPCHF", "GBPJPY",
    "AUDNZD", "AUDUSD", "AUDCAD", "AUDCHF", "AUDJPY",
    "NZDUSD", "NZDCAD", "NZDCHF", "NZDJPY",
    "USDCAD", "USDCHF", "USDJPY",
    "CADCHF", "CADJPY", "CHFJPY",
)

SOURCES = {
    "USD": {
        "name": "US Treasury 2Y",
        "url": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView?type=daily_treasury_yield_curve",
        "cadence": "daily",
        "route": "arquivos anuais oficiais do Treasury",
        "current_bdays": 3,
        "delayed_bdays": 7,
    },
    "EUR": {
        "name": "ECB AAA 2Y spot",
        "url": "https://data.ecb.europa.eu/data/datasets/YC",
        "cadence": "daily",
        "route": "API oficial ECB",
        "current_bdays": 3,
        "delayed_bdays": 7,
    },
    "GBP": {
        "name": "Bank of England 2Y spot curve",
        "url": "https://www.bankofengland.co.uk/statistics/yield-curves",
        "cadence": "daily",
        "route": "arquivo histórico oficial BoE",
        "current_bdays": 3,
        "delayed_bdays": 7,
    },
    "JPY": {
        "name": "Japan MoF JGB 2Y",
        "url": "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/",
        "cadence": "daily",
        "route": "arquivo histórico oficial MoF",
        "current_bdays": 3,
        "delayed_bdays": 7,
    },
    "AUD": {
        "name": "RBA F2 Australian Government 2Y",
        "url": "https://www.rba.gov.au/statistics/tables/",
        "cadence": "weekly",
        "route": "RBA histórico direto + atual via DBnomics",
        "current_bdays": 7,
        "delayed_bdays": 12,
    },
    "CAD": {
        "name": "Bank of Canada benchmark 2Y",
        "url": "https://www.bankofcanada.ca/rates/interest-rates/lookup-bond-yields/",
        "cadence": "daily",
        "route": "API Valet oficial",
        "current_bdays": 3,
        "delayed_bdays": 7,
    },
    "NZD": {
        "name": "RBNZ B2 Government 2Y",
        "url": "https://www.rbnz.govt.nz/statistics/series/exchange-and-interest-rates/wholesale-interest-rates",
        "cadence": "daily",
        "route": "arquivos oficiais histórico + atual",
        "current_bdays": 3,
        "delayed_bdays": 7,
    },
    "CHF": {
        "name": "SNB Confederation NSS spot rate 2Y",
        "url": "https://data.snb.ch/en/warehouse/SNB1A/cube/SNB1A@SNB.NSS.KZS.EID",
        "cadence": "daily (business days), published D+1 ~08:30 CET",
        "route": "API warehouse oficial",
        "current_bdays": 3,
        "delayed_bdays": 7,
    },
}

DOWNLOADS = {
    "AUD": (
        "https://api.db.nomics.world/v22/series/RBA/F2/FCMYGBAG2D?observations=1",
        RAW_DIR / "aud_dbnomics_rba_f2.json",
        10_000,
        18,
    ),
    "CAD": (
        "https://www.bankofcanada.ca/valet/observations/BD.CDN.2YR.DQ.YLD/json?start_date=2002-01-01",
        RAW_DIR / "cad_boc_2y.json",
        50_000,
        18,
    ),
    "NZD": (
        "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/series/b/b2/hb2-daily-close.xlsx?download=1",
        RAW_DIR / "nzd_rbnz_b2.xlsx",
        50_000,
        18,
    ),
    # Cube WAREHOUSE do SNB: mesma curva, mas publicada DIARIAMENTE (D+1 ~08:30 CET),
    # enquanto o cube publico rendeiduebd so sai em lote mensal (ate 31 dias de atraso).
    # Os valores conferem exatamente entre as duas fontes (ex.: 2026-07-31 = 0,089).
    # Pegadinha da API: no path o "@" do cubeId vira "."; sem fromDate ela devolve
    # apenas os ultimos ~30 dias.
    "CHF": (
        "https://data.snb.ch/api/warehouse/cube/SNB1A.SNB.NSS.KZS.EID/data/csv/en"
        "?dimSel=LAUFZEIT(J02M0),ZEITPUNKT(A1100),frequency(P1D_L),AGGREGATIONSMETHODE(ZZ)"
        "&fromDate=1988-01-01&toDate=2036-12-31",
        RAW_DIR / "chf_snb_nss_2y.csv",
        150_000,
        72,
    ),
}

HISTORY_DOWNLOADS = {
    "EUR_HISTORY": (
        "https://data-api.ecb.europa.eu/service/data/YC/B.U2.EUR.4F.G_N_A.SV_C_YM.SR_2Y?startPeriod=2004-09-06&format=csvdata",
        RAW_DIR / "eur_ecb_2y.csv",
        1_000_000,
        18,
    ),
    "GBP_HISTORY": (
        "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/glcnominalddata.zip",
        RAW_DIR / "gbp_glcnominalddata.zip",
        10_000_000,
        168,
    ),
    "JPY_HISTORY": (
        "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/historical/jgbcme_all.csv",
        RAW_DIR / "jpy_mof_all.csv",
        500_000,
        168,
    ),
    # 31-ago-2026: o arquivo historico do MoF termina no mes ANTERIOR. Este e o do mes
    # corrente, e e ele que fecha o buraco que impedia medir USDJPY, AUDJPY e companhia.
    # 04-set-2026: TTL de 18 h deixou o arquivo parado em 28/08 por uma semana — a cadeia roda
    # 2x/dia e a segunda rodada sempre pulava; a guarda de frescor ficou vermelha 2x/dia e o
    # e-mail de falha chegou ao Eduardo. O arquivo tem 20 KB e o MOF publica todo dia util as
    # 23:30 GMT (conferido: fonte com dado de 03/09 enquanto o repo tinha 28/08). TTL de 1 h:
    # toda rodada baixa de novo. Custo zero, e a guarda volta a medir a FONTE, nao o cache.
    "JPY_CURRENT": (
        "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/jgbcme.csv",
        RAW_DIR / "jpy_mof_mes.csv",
        800,
        1,
    ),
    "AUD_HISTORY": (
        "https://www.rba.gov.au/statistics/tables/xls-hist/f02dhist.xls",
        RAW_DIR / "aud_rba_f02dhist.xls",
        500_000,
        24 * 3650,
    ),
    "NZD_HISTORY": (
        "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/series/b/b2/hb2-daily-close-1985-2017.xlsx?download=1",
        RAW_DIR / "nzd_rbnz_b2_daily_close_1985_2017.xlsx",
        300_000,
        24 * 3650,
    ),
}

FX_DOWNLOADS = {
    f"FX_{currency}": (
        f"https://data-api.ecb.europa.eu/service/data/EXR/D.{currency}.EUR.SP00.A?startPeriod=2002-01-01&format=csvdata",
        RAW_DIR / f"fx_ecb_{currency.lower()}eur.csv",
        100_000,
        18,
    )
    for currency in CURRENCIES
    if currency != "EUR"
}

NEWS_DOWNLOADS = {
    "NEWS_WEEK": (
        "https://nfs.faireconomy.media/ff_calendar_thisweek.json",
        RAW_DIR / "ff_calendar_thisweek.json",
        1_000,
        2,
    ),
}


@dataclass
class DownloadResult:
    currency: str
    state: str
    detail: str


def log(message: str) -> None:
    print(message, flush=True)


def parse_iso(value: str) -> date:
    return date.fromisoformat(value[:10])


def add_business_day(day: date, count: int = 1) -> date:
    out = day
    remaining = count
    while remaining:
        out += timedelta(days=1)
        if out.weekday() < 5:
            remaining -= 1
    return out


def business_days(start: date, end: date) -> list[date]:
    days: list[date] = []
    cursor = start
    while cursor <= end:
        if cursor.weekday() < 5:
            days.append(cursor)
        cursor += timedelta(days=1)
    return days


def business_age(last_day: date, today: date) -> int:
    if last_day >= today:
        return 0
    return sum(1 for d in business_days(last_day + timedelta(days=1), today))


def shift_pit(series: dict[date, float | None]) -> dict[date, float | None]:
    """Lag conservador de +1 dia util, igual ao pipeline FUND oficial."""
    return {add_business_day(d): value for d, value in series.items()}


def download_resource(
    key: str,
    specification: tuple[str, Path, int, int],
    force: bool,
    referer: str,
) -> DownloadResult:
    url, path, min_bytes, ttl_hours = specification
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and not force:
        # ⚠️ 31-ago-2026 — O BUG QUE CONGELOU TODAS AS FONTES POR UMA SEMANA.
        # Isto usava path.stat().st_mtime. Na maquina local funcionava; na nuvem, nao:
        # o actions/checkout REESCREVE todo arquivo a cada execucao, entao o mtime e
        # sempre "agora", o TTL dava sempre "fresco" e o download NUNCA acontecia.
        # O carimbo abaixo grava a hora REAL do download num arquivo que e commitado,
        # entao sobrevive ao checkout.
        try:
            import frescor
            idade_h = frescor.horas_desde_download(path.name)
        except Exception:
            idade_h = None
        if idade_h is None:
            idade_h = (datetime.now().timestamp() - path.stat().st_mtime) / 3600.0
        if idade_h < ttl_hours and path.stat().st_size >= min_bytes:
            return DownloadResult(key, "CACHE", f"{path.stat().st_size:,} bytes")

    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/139 Safari/537.36",
            "Accept": "application/json,text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            "Accept-Language": "en-US,en;q=0.9,pt-BR;q=0.8",
            "Referer": referer,
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=180) as response:
            payload = response.read()
        if len(payload) < min_bytes:
            raise RuntimeError(f"resposta suspeita: {len(payload)} bytes")
        path.write_bytes(payload)
        try:
            import frescor; frescor.grava_carimbo(path.name)
        except Exception:
            pass
        return DownloadResult(key, "BAIXADO", f"{len(payload):,} bytes")
    except Exception as exc:  # algumas fontes (RBNZ) recusam urllib, mas aceitam curl
        temp = path.with_suffix(path.suffix + ".download")
        try:
            completed = subprocess.run(
                [
                    # 31-ago-2026: era "curl.exe" — nome do binario no Windows. No runner do
                    # GitHub (Linux) o arquivo nao existe, entao o fallback falhava
                    # calado e o cache velho era mantido.
                    "curl", "-L", "--fail", "--retry", "2",
                    "--connect-timeout", "30", "-A",
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/139 Safari/537.36",
                    "-e", referer, "-o", str(temp), url,
                ],
                check=False,
                capture_output=True,
                timeout=240,
            )
            if completed.returncode == 0 and temp.exists() and temp.stat().st_size >= min_bytes:
                size = temp.stat().st_size
                temp.replace(path)
                try:
                    import frescor; frescor.grava_carimbo(path.name)
                except Exception:
                    pass
                return DownloadResult(key, "BAIXADO_CURL", f"{size:,} bytes")
        except Exception:
            pass
        finally:
            if temp.exists():
                temp.unlink()
        if path.exists() and path.stat().st_size >= min_bytes:
            return DownloadResult(key, "CACHE_APOS_FALHA", str(exc))
        return DownloadResult(key, "FALHOU", str(exc))


def download(currency: str, force: bool) -> DownloadResult:
    return download_resource(currency, DOWNLOADS[currency], force, SOURCES[currency]["url"])


def refresh_usd_history(force: bool, offline: bool) -> list[DownloadResult]:
    UST_DIR.mkdir(parents=True, exist_ok=True)
    current_year = datetime.now(ZoneInfo("America/Sao_Paulo")).year
    results: list[DownloadResult] = []
    for year in range(CALENDAR_START.year, current_year + 1):
        key = f"USD_{year}"
        path = UST_DIR / f"ust_cmt_{year}.csv"
        if offline:
            results.append(DownloadResult(key, "OFFLINE", "cache local"))
            continue
        # ⚠️ 31-ago-2026: esta URL estava MALFORMADA e devolvia HTTP 406 ha quem sabe
        # quanto tempo, em silencio — o download caia em CACHE_APOS_FALHA e ninguem via.
        # Os dois pedacos estavam colados: ".../TextView?type=daily_treasury_yield_curve"
        # + "daily-treasury-rates.csv/..." virava "yield_curvedaily-treasury-rates.csv",
        # com DOIS pontos de interrogacao na mesma URL. O Treasury respondia 406.
        url = (
            "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/"
            f"daily-treasury-rates.csv/{year}/all"
            f"?type=daily_treasury_yield_curve&field_tdr_date_value={year}&page&_format=csv"
        )
        # Anos encerrados sao imutaveis. --force atualiza somente o ano corrente.
        results.append(download_resource(
            key, (url, path, 4_000, 18 if year == current_year else 24 * 3650),
            force and year == current_year,
            "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/TextView",
        ))
    return results


def refresh_downloads(force: bool, offline: bool) -> list[DownloadResult]:
    usd_keys = [f"USD_{year}" for year in range(CALENDAR_START.year, date.today().year + 1)]
    all_keys = [*DOWNLOADS, *HISTORY_DOWNLOADS, *FX_DOWNLOADS, *NEWS_DOWNLOADS, *usd_keys]
    if offline:
        return [DownloadResult(key, "OFFLINE", "cache local") for key in all_keys]
    results = [download(currency, force) for currency in DOWNLOADS]
    history_referers = {
        "EUR_HISTORY": SOURCES["EUR"]["url"],
        "GBP_HISTORY": SOURCES["GBP"]["url"],
        "JPY_HISTORY": SOURCES["JPY"]["url"],
        "JPY_CURRENT": SOURCES["JPY"]["url"],
        "AUD_HISTORY": "https://www.rba.gov.au/statistics/historical-data.html",
        "NZD_HISTORY": SOURCES["NZD"]["url"],
    }
    results.extend(
        download_resource(
            key, spec,
            force and key in ("EUR_HISTORY", "GBP_HISTORY", "JPY_HISTORY"),
            history_referers[key],
        )
        for key, spec in HISTORY_DOWNLOADS.items()
    )
    results.extend(
        download_resource(key, spec, force, "https://data.ecb.europa.eu/")
        for key, spec in FX_DOWNLOADS.items()
    )
    results.extend(
        download_resource(key, spec, force, "https://www.forexfactory.com/calendar")
        for key, spec in NEWS_DOWNLOADS.items()
    )
    results.extend(refresh_usd_history(force, offline=False))
    return results


def read_usd_history() -> dict[date, float]:
    output: dict[date, float] = {}
    for path in sorted(UST_DIR.glob("ust_cmt_*.csv")):
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                raw = (row.get("2 Yr") or "").strip()
                if not raw:
                    continue
                try:
                    day = datetime.strptime(row["Date"], "%m/%d/%Y").date()
                    output[day] = float(raw)
                except (KeyError, ValueError):
                    continue
    return output


def read_eur_history() -> dict[date, float]:
    path = HISTORY_DOWNLOADS["EUR_HISTORY"][1]
    output: dict[date, float] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            raw = (row.get("OBS_VALUE") or "").strip()
            if raw:
                output[parse_iso(row["TIME_PERIOD"])] = float(raw)
    return output


def read_gbp_history() -> dict[date, float]:
    path = HISTORY_DOWNLOADS["GBP_HISTORY"][1]
    output: dict[date, float] = {}
    with zipfile.ZipFile(path) as archive:
        members = sorted(name for name in archive.namelist() if name.lower().endswith(".xlsx"))
        for member in members:
            workbook = load_workbook(io.BytesIO(archive.read(member)), read_only=True, data_only=True)
            sheet_name = next(
                (name for name in workbook.sheetnames if "spot curve" in name.lower() and name.lstrip().startswith("4.")),
                None,
            )
            if sheet_name is None:
                continue
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            maturity_row: tuple | None = None
            for _ in range(10):
                candidate = next(rows, None)
                if candidate is None:
                    break
                if any(isinstance(value, (int, float)) and abs(float(value) - 2.0) < 1e-9 for value in candidate[1:]):
                    maturity_row = candidate
                    break
            if maturity_row is None:
                continue
            target_column = next(
                index for index, value in enumerate(maturity_row)
                if isinstance(value, (int, float)) and abs(float(value) - 2.0) < 1e-9
            )
            for row in rows:
                if not row or not isinstance(row[0], (date, datetime)) or target_column >= len(row):
                    continue
                raw = row[target_column]
                if isinstance(raw, (int, float)):
                    day = row[0].date() if isinstance(row[0], datetime) else row[0]
                    output[day] = float(raw)
            workbook.close()
    return output


def _le_jgb(path, pular: int) -> dict[date, float]:
    """Le um CSV do MoF. Os dois tem o mesmo corpo; muda so quantas linhas de aviso pular."""
    out: dict[date, float] = {}
    if not path.exists():
        return out
    with path.open("r", encoding="cp932", errors="replace", newline="") as handle:
        for _ in range(pular):
            next(handle, None)
        for row in csv.DictReader(handle):
            raw = (row.get("2Y") or "").strip()
            if raw in ("", "-"):
                continue
            try:
                out[datetime.strptime(row["Date"].strip(), "%Y/%m/%d").date()] = float(raw)
            except (KeyError, ValueError, AttributeError):
                continue
    return out


def read_jpy_history() -> dict[date, float]:
    """Historico + mes corrente.

    O arquivo historico do MoF termina no mes anterior — em 31/ago/2026 ele parava em
    31/07, e isso deixava o JPY sem observacao por quatro semanas, o que impedia medir
    qualquer par com iene. O arquivo do mes corrente cobre o resto.
    """
    base = _le_jgb(HISTORY_DOWNLOADS["JPY_HISTORY"][1], 1)
    mes = _le_jgb(HISTORY_DOWNLOADS["JPY_CURRENT"][1], 1)
    base.update(mes)          # o mes corrente tem prioridade onde houver sobreposicao
    return base


def read_core() -> dict[str, dict[date, float | None]]:
    output: dict[str, dict[date, float | None]] = {
        "USD": read_usd_history(),
        "EUR": read_eur_history(),
        "GBP": read_gbp_history(),
        "JPY": read_jpy_history(),
    }
    # O consolidado existente acrescenta as observacoes mais recentes e
    # preserva a reproducao exata do snapshot ja auditado.
    if CORE_YIELDS.exists():
        with CORE_YIELDS.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                day = parse_iso(row["date"])
                for currency in output:
                    raw = (row.get(currency.lower()) or "").strip()
                    if raw:
                        output[currency][day] = float(raw)
                    elif day not in output[currency]:
                        output[currency][day] = None
    return output


def read_aud() -> dict[date, float]:
    historical_path = HISTORY_DOWNLOADS["AUD_HISTORY"][1]
    _, rows, date_mode = read_biff_sheet(historical_path, ("Data",))
    mnemonic_row = next(
        row for row in rows
        if any(str(value).strip() == "FCMYGBAG2D" for value in row if value is not None)
    )
    target_column = next(
        index for index, value in enumerate(mnemonic_row)
        if str(value).strip() == "FCMYGBAG2D"
    )
    excel_epoch = date(1904, 1, 1) if date_mode else date(1899, 12, 30)
    output: dict[date, float] = {}
    for row in rows:
        if not row or not isinstance(row[0], (int, float)) or target_column >= len(row):
            continue
        raw = row[target_column]
        if isinstance(raw, (int, float)) and row[0] > 20_000:
            output[excel_epoch + timedelta(days=int(row[0]))] = float(raw)

    path = DOWNLOADS["AUD"][1]
    payload = json.loads(path.read_text(encoding="utf-8"))
    doc = payload["series"]["docs"][0]
    for period, value in zip(doc["period"], doc["value"]):
        if value is None:
            continue
        try:
            output[parse_iso(period)] = float(str(value).replace(",", "."))
        except ValueError:
            continue
    return output


def read_cad() -> dict[date, float]:
    path = DOWNLOADS["CAD"][1]
    payload = json.loads(path.read_text(encoding="utf-8"))
    output: dict[date, float] = {}
    for row in payload["observations"]:
        raw = row.get("BD.CDN.2YR.DQ.YLD", {}).get("v")
        if raw not in (None, ""):
            output[parse_iso(row["d"])] = float(raw)
    return output


def column_number(reference: str) -> int:
    letters = re.match(r"[A-Z]+", reference)
    if not letters:
        raise ValueError(f"celula XLSX invalida: {reference}")
    number = 0
    for char in letters.group(0):
        number = number * 26 + ord(char) - 64
    return number


def read_nzd_xlsx(path: Path) -> dict[date, float]:
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as archive:
        shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
        shared = ["".join(node.itertext()) for node in shared_root.findall("x:si", namespace)]
        sheet = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))

    rows: dict[int, dict[int, str]] = {}
    for row in sheet.findall(".//x:sheetData/x:row", namespace):
        row_number = int(row.attrib["r"])
        cells: dict[int, str] = {}
        for cell in row.findall("x:c", namespace):
            ref = cell.attrib["r"]
            value_node = cell.find("x:v", namespace)
            raw = "" if value_node is None or value_node.text is None else value_node.text
            if cell.attrib.get("t") == "s" and raw:
                raw = shared[int(raw)]
            cells[column_number(ref)] = raw
        rows[row_number] = cells

    series_row = rows.get(5, {})
    target_column = next(
        (column for column, value in series_row.items() if value == "INM.DG102.NZZCF"),
        None,
    )
    if target_column is None:
        raise RuntimeError("RBNZ: coluna Government 2Y nao encontrada")

    excel_epoch = date(1899, 12, 30)
    output: dict[date, float] = {}
    for row_number in sorted(number for number in rows if number >= 6):
        cells = rows[row_number]
        serial = cells.get(1, "")
        raw = cells.get(target_column, "")
        if not serial or not raw:
            continue
        try:
            day = excel_epoch + timedelta(days=int(float(serial)))
            output[day] = float(raw)
        except ValueError:
            continue
    return output


def read_nzd() -> dict[date, float]:
    output = read_nzd_xlsx(HISTORY_DOWNLOADS["NZD_HISTORY"][1])
    output.update(read_nzd_xlsx(DOWNLOADS["NZD"][1]))
    return output


def read_chf() -> dict[date, float]:
    path = DOWNLOADS["CHF"][1]
    output: dict[date, float] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        started = False
        for line in handle:
            if not started:
                if line.lstrip("\ufeff").startswith('"Date";'):
                    started = True
                continue
            row = next(csv.reader([line], delimiter=";"))
            # warehouse NSS: Date;LAUFZEIT;ZEITPUNKT;frequency;AGGREGATIONSMETHODE;Value
            if len(row) >= 6 and row[1] == "J02M0" and row[5]:
                output[parse_iso(row[0])] = float(row[5])
            # compatibilidade com o cube antigo: Date;D0;D1;Value
            elif len(row) == 4 and row[1] == "CHF" and row[2] == "2J" and row[3]:
                output[parse_iso(row[0])] = float(row[3])
    return output


def read_all_yields() -> dict[str, dict[date, float | None]]:
    series = read_core()
    series.update({"AUD": read_aud(), "CAD": read_cad(), "NZD": read_nzd(), "CHF": read_chf()})
    # NOWCAST (24/ago/2026): as fontes oficiais chegam com 2-4 dias de atraso,
    # mas o yield 2y e um preco negociado ao vivo. tv_yields_nowcast.py grava o
    # valor de HOJE (TradingView TVC); aqui ele entra como ponto adicional SO
    # quando a serie oficial ainda nao tem o dia — o oficial continua sendo a
    # espinha dorsal historica. Nowcast com mais de 24h e ignorado.
    try:
        nc = json.loads((ROOT / "data" / "raw" / "tv_nowcast.json").read_text(encoding="utf-8"))
        fetched = datetime.fromisoformat(nc["fetched_at"])
        idade_h = (datetime.now(timezone.utc) - fetched).total_seconds() / 3600.0
        dia_nc = parse_iso(nc["date"])
        if idade_h <= 24.0:
            for moeda, item in (nc.get("yields") or {}).items():
                s = series.get(moeda)
                if s is None:
                    continue
                datas = [d for d, v in s.items() if v is not None]
                if datas and max(datas) < dia_nc:
                    s[dia_nc] = float(item["value"])
    except FileNotFoundError:
        pass
    except Exception as e:
        print(f"nowcast ignorado ({e})")
    return series


def read_fx_rates() -> dict[str, dict[date, float]]:
    """Le referencias oficiais do ECB em unidades da moeda por 1 EUR."""
    rates: dict[str, dict[date, float]] = {"EUR": {}}
    all_dates: set[date] = set()
    for key, (_, path, _, _) in FX_DOWNLOADS.items():
        currency = key.removeprefix("FX_")
        series: dict[date, float] = {}
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                raw = (row.get("OBS_VALUE") or "").strip()
                if raw:
                    day = parse_iso(row["TIME_PERIOD"])
                    series[day] = float(raw)
                    all_dates.add(day)
        if not series:
            raise RuntimeError(f"ECB FX {currency}: serie vazia")
        rates[currency] = series
    rates["EUR"] = {day: 1.0 for day in all_dates}
    return rates


def derive_pair_prices(pair: str, rates: dict[str, dict[date, float]]) -> dict[date, float]:
    """Converte referencias EUR em preco quote por uma unidade de base."""
    base = rates[pair[:3]]
    quote = rates[pair[3:]]
    calendar = sorted(set(base) | set(quote))
    base_values = ffill(base, calendar, FFILL_LIMIT)
    quote_values = ffill(quote, calendar, FFILL_LIMIT)
    output: dict[date, float] = {}
    for day, base_value, quote_value in zip(calendar, base_values, quote_values):
        if base_value not in (None, 0.0) and quote_value is not None:
            output[day] = quote_value / base_value
    return output


def event_scenarios(title: str, currency: str) -> dict[str, str]:
    lowered = title.lower()
    speech_words = ("speaks", "speech", "minutes", "press conference", "report")
    inverse_words = ("unemployment", "claims", "claimant", "jobless")
    inflation_words = ("cpi", "ppi", "inflation", "price index", "wage")
    rate_words = ("interest rate", "cash rate", "refinancing rate", "bank rate", "rate statement")
    growth_words = (
        "gdp", "employment change", "payroll", "retail sales", "pmi", "production",
        "confidence", "sentiment", "trade balance", "orders", "manufacturing",
    )

    if any(word in lowered for word in speech_words):
        category = "COMUNICACAO"
        above = f"A more hawkish tone tends to lift expected rates and strengthen {currency}."
        below = f"A more dovish tone tends to lower expected rates and weaken {currency}."
    elif any(word in lowered for word in inverse_words):
        category = "TRABALHO_INVERSO"
        above = f"Acima do consenso indica piora no trabalho e tende a enfraquecer {currency}."
        below = f"Abaixo do consenso indica trabalho mais firme e tende a fortalecer {currency}."
    elif any(word in lowered for word in rate_words):
        category = "JUROS"
        above = f"Taxa ou orientação acima do esperado tende a fortalecer {currency}."
        below = f"Taxa ou orientação abaixo do esperado tende a enfraquecer {currency}."
    elif any(word in lowered for word in inflation_words):
        category = "INFLACAO"
        above = f"Inflação acima do consenso tende a elevar juros esperados e fortalecer {currency}."
        below = f"Inflação abaixo do consenso tende a reduzir juros esperados e enfraquecer {currency}."
    elif any(word in lowered for word in growth_words):
        category = "ATIVIDADE"
        above = f"Activity above consensus tends to strengthen {currency}."
        below = f"Activity below consensus tends to weaken {currency}."
    else:
        category = "CONTEXTO"
        above = f"A reading better than consensus may favour {currency}; confirm the reaction in yields."
        below = f"A reading worse than consensus may pressure {currency}; confirm the reaction in yields."
    return {"category": category, "above": above, "below": below}


def read_news(pairs: list[dict]) -> list[dict]:
    path = NEWS_DOWNLOADS["NEWS_WEEK"][1]
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    now = datetime.now(ZoneInfo("America/Sao_Paulo"))
    events: list[dict] = []
    for item in payload:
        currency = str(item.get("country", "")).upper()
        if currency not in CURRENCIES:
            continue
        try:
            event_time = datetime.fromisoformat(item["date"])
        except (KeyError, ValueError):
            continue
        if event_time.astimezone(ZoneInfo("America/Sao_Paulo")) < now - timedelta(minutes=5):
            continue
        scenarios = event_scenarios(str(item.get("title", "")), currency)
        related = [
            pair["pair"]
            for pair in sorted(pairs, key=lambda row: abs(row["fund"]), reverse=True)
            if currency in (pair["base"], pair["quote"])
        ][:4]
        events.append({
            "title": item.get("title", ""),
            "currency": currency,
            "date": event_time.isoformat(),
            "impact": item.get("impact", "Low"),
            "forecast": item.get("forecast", ""),
            "previous": item.get("previous", ""),
            "category": scenarios["category"],
            "above_scenario": scenarios["above"],
            "below_scenario": scenarios["below"],
            "related_pairs": related,
        })
    impact_rank = {"High": 0, "Medium": 1, "Low": 2, "Holiday": 3}
    events.sort(key=lambda row: (row["date"], impact_rank.get(row["impact"], 9)))
    return events[:40]


def ffill(series: dict[date, float | None], calendar: Iterable[date], limit: int) -> list[float | None]:
    values: list[float | None] = []
    last: float | None = None
    gap = limit + 1
    for day in calendar:
        if day in series and series[day] is not None:
            last = series[day]
            gap = 0
            values.append(last)
        else:
            gap += 1
            values.append(last if last is not None and gap <= limit else None)
    return values


def classify(score: float | None) -> str:
    if score is None:
        return "SEM_DADO"
    if score >= 60:
        return "STRONG_BULL"
    if score >= 25:
        return "BULL"
    if score <= -60:
        return "STRONG_BEAR"
    if score <= -25:
        return "BEAR"
    return "NEUTRAL"


RANK = {"STRONG_BEAR": -2, "BEAR": -1, "NEUTRAL": 0, "BULL": 1, "STRONG_BULL": 2}


def strength_change(previous: str, current: str) -> tuple[str | None, str | None]:
    if previous not in RANK or current not in RANK or previous == current:
        return None, None
    if previous in ("BULL", "STRONG_BULL") and RANK[current] < RANK[previous]:
        return "SAIR_LONG", "FUND lost a band of buying strength"
    if previous in ("BEAR", "STRONG_BEAR") and RANK[current] > RANK[previous]:
        return "SAIR_SHORT", "FUND lost a band of selling strength"
    if abs(RANK[current]) > abs(RANK[previous]):
        return "FORTALECEU", "FUND gained a band in the current direction"
    return "MUDOU_FAIXA", "FUND mudou de faixa"


def compute_pair(pair: str, base_series: dict[date, float | None], quote_series: dict[date, float | None]) -> dict:
    # A implementacao oficial usa o indice-uniao das duas series; shift(20),
    # rolling(252) e ffill(limit=5) contam linhas desse indice, nao um calendario
    # sintetico. Preservar isso e necessario para reproduzir exatamente o V0.1.
    calendar = sorted(set(base_series) | set(quote_series))
    base_values = ffill(base_series, calendar, FFILL_LIMIT)
    quote_values = ffill(quote_series, calendar, FFILL_LIMIT)
    spread: list[float | None] = []
    raw: list[float | None] = []
    fund: list[float | None] = []

    for index, (base_value, quote_value) in enumerate(zip(base_values, quote_values)):
        current_spread = None if base_value is None or quote_value is None else base_value - quote_value
        spread.append(current_spread)
        old_spread = spread[index - LOOKBACK] if index >= LOOKBACK else None
        current_raw = None if current_spread is None or old_spread is None else current_spread - old_spread
        raw.append(current_raw)

        history = [value for value in raw[max(0, index - NORM_WINDOW):index] if value is not None]
        if current_raw is None or len(history) < MIN_HISTORY:
            fund.append(None)
            continue
        sigma = statistics.stdev(history)
        if sigma == 0:
            fund.append(None)
            continue
        z_score = (current_raw - statistics.mean(history)) / sigma
        fund.append(max(-100.0, min(100.0, 100.0 * math.tanh(z_score / 2.0))))

    valid_indexes = [index for index, value in enumerate(fund) if value is not None]
    if not valid_indexes:
        raise RuntimeError(f"{pair}: historico insuficiente")
    last_index = valid_indexes[-1]
    previous_index = valid_indexes[-2] if len(valid_indexes) >= 2 else last_index
    latest_score = fund[last_index]
    assert latest_score is not None
    current_label = classify(latest_score)
    previous_label = classify(fund[previous_index])
    alert, alert_detail = strength_change(previous_label, current_label)
    five_back = valid_indexes[-6] if len(valid_indexes) >= 6 else valid_indexes[0]

    history = []
    for index in valid_indexes[-260:]:
        value = fund[index]
        history.append({
            "date": calendar[index].isoformat(),
            "fund": round(value, 2) if value is not None else None,
            "spread": round(spread[index], 4) if spread[index] is not None else None,
            "momentum20": round(raw[index], 4) if raw[index] is not None else None,
        })
    full_fund_history = [
        {
            "date": calendar[index].isoformat(),
            "fund": round(fund[index], 6),
            "strength": classify(fund[index]),
        }
        for index in valid_indexes
        if fund[index] is not None
    ]

    # ha quantos pregoes o FUND esta na MESMA faixa? O backtest de maturidade
    # (SWING_FOREX_PREREG_MATURIDADE_FUND.md) mostrou PF ~1,0 abaixo de 4 pregoes
    # e PF 4,3 a partir dai — o painel usa isso para ligar a vigilancia.
    dias_na_faixa = 0
    for item in reversed(full_fund_history):
        if item["strength"] == current_label:
            dias_na_faixa += 1
        else:
            break

    return {
        "pair": pair,
        "base": pair[:3],
        "quote": pair[3:],
        "as_of": calendar[last_index].isoformat(),
        "fund": round(latest_score, 2),
        "strength": current_label,
        "days_in_band": dias_na_faixa,
        "previous_strength": previous_label,
        "change_5d": round(latest_score - (fund[five_back] or 0.0), 2),
        "spread": round(spread[last_index], 4) if spread[last_index] is not None else None,
        "momentum20": round(raw[last_index], 4) if raw[last_index] is not None else None,
        "yield_base": round(base_values[last_index], 4) if base_values[last_index] is not None else None,
        "yield_quote": round(quote_values[last_index], 4) if quote_values[last_index] is not None else None,
        "exit_alert": alert,
        "exit_detail": alert_detail,
        "history": history,
        "_fund_history": full_fund_history,
    }


def trade_return_pct(pair: str, side: int, entry: float, exit_price: float, cost_pips: float) -> tuple[float, float]:
    gross = side * ((exit_price - entry) / entry) * 100.0
    pip_size = 0.01 if pair.endswith("JPY") else 0.0001
    cost_pct = (cost_pips * pip_size / entry) * 100.0
    return gross - cost_pct, cost_pct


def backtest_pair(
    pair: str,
    fund_history: list[dict],
    prices: dict[date, float],
    start: date,
    end: date,
    cost_pips: float = DEFAULT_COST_PIPS,
) -> dict:
    """Backtest causal do motor FUND: sinal em D, execucao no proximo fix ECB."""
    price_dates = sorted(day for day in prices if start <= day <= end)
    if len(price_dates) < 2:
        return summarize_backtest(pair, [], start, end, cost_pips)

    def execution_after(signal_day: date) -> tuple[date, float] | None:
        index = bisect.bisect_right(price_dates, signal_day)
        if index >= len(price_dates):
            return None
        day = price_dates[index]
        return day, prices[day]

    rows = [
        (parse_iso(row["date"]), float(row["fund"]), row["strength"])
        for row in fund_history
        if parse_iso(row["date"]) <= end
    ]
    position: dict | None = None
    trades: list[dict] = []
    previous_strength: str | None = None
    started = False

    for signal_day, score, current_strength in rows:
        if signal_day < start:
            continue
        if not started:
            previous_strength = "NEUTRAL"
            started = True
        execution = execution_after(signal_day)

        if position is not None and previous_strength is not None and execution is not None:
            alert, _ = strength_change(previous_strength, current_strength)
            exit_matches = (
                position["side"] == 1 and alert == "SAIR_LONG"
            ) or (
                position["side"] == -1 and alert == "SAIR_SHORT"
            )
            if exit_matches:
                exit_day, exit_price = execution
                net_return, cost_pct = trade_return_pct(
                    pair, position["side"], position["entry_price"], exit_price, cost_pips
                )
                trades.append({
                    "pair": pair,
                    "side": "LONG" if position["side"] == 1 else "SHORT",
                    "signal_date": position["signal_date"].isoformat(),
                    "entry_date": position["entry_date"].isoformat(),
                    "exit_signal_date": signal_day.isoformat(),
                    "exit_date": exit_day.isoformat(),
                    "entry_price": round(position["entry_price"], 6),
                    "exit_price": round(exit_price, 6),
                    "fund_entry": round(position["fund_entry"], 2),
                    "entry_strength": position["entry_strength"],
                    "exit_strength": current_strength,
                    "return_pct": round(net_return, 4),
                    "cost_pct": round(cost_pct, 4),
                    "holding_days": (exit_day - position["entry_date"]).days,
                    "reason": "PERDEU_FAIXA_FUND",
                })
                position = None

        if position is None and execution is not None and previous_strength is not None:
            current_bull = current_strength in ("BULL", "STRONG_BULL")
            previous_bull = previous_strength in ("BULL", "STRONG_BULL")
            current_bear = current_strength in ("BEAR", "STRONG_BEAR")
            previous_bear = previous_strength in ("BEAR", "STRONG_BEAR")
            side = 1 if current_bull and not previous_bull else -1 if current_bear and not previous_bear else 0
            if side:
                entry_day, entry_price = execution
                position = {
                    "side": side,
                    "signal_date": signal_day,
                    "entry_date": entry_day,
                    "entry_price": entry_price,
                    "fund_entry": score,
                    "entry_strength": current_strength,
                }
        previous_strength = current_strength

    if position is not None:
        last_day = price_dates[-1]
        if last_day >= position["entry_date"]:
            exit_price = prices[last_day]
            net_return, cost_pct = trade_return_pct(
                pair, position["side"], position["entry_price"], exit_price, cost_pips
            )
            trades.append({
                "pair": pair,
                "side": "LONG" if position["side"] == 1 else "SHORT",
                "signal_date": position["signal_date"].isoformat(),
                "entry_date": position["entry_date"].isoformat(),
                "exit_signal_date": None,
                "exit_date": last_day.isoformat(),
                "entry_price": round(position["entry_price"], 6),
                "exit_price": round(exit_price, 6),
                "fund_entry": round(position["fund_entry"], 2),
                "entry_strength": position["entry_strength"],
                "exit_strength": None,
                "return_pct": round(net_return, 4),
                "cost_pct": round(cost_pct, 4),
                "holding_days": (last_day - position["entry_date"]).days,
                "reason": "FIM_DA_JANELA",
            })
    return summarize_backtest(pair, trades, start, end, cost_pips)


def summarize_backtest(pair: str, trades: list[dict], start: date, end: date, cost_pips: float) -> dict:
    returns = [float(trade["return_pct"]) for trade in trades]
    wins = [value for value in returns if value > 0]
    losses = [value for value in returns if value < 0]
    gross_profit = sum(wins)
    gross_loss = abs(sum(losses))
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else None
    equity = 1.0
    peak = 1.0
    max_drawdown = 0.0
    curve = [{"date": start.isoformat(), "equity": 100.0}]
    for trade in trades:
        equity *= max(0.0001, 1.0 + float(trade["return_pct"]) / 100.0)
        peak = max(peak, equity)
        max_drawdown = max(max_drawdown, (peak - equity) / peak * 100.0)
        curve.append({"date": trade["exit_date"], "equity": round(equity * 100.0, 4)})
    hold_values = [trade["holding_days"] for trade in trades]
    summary = {
        "pair": pair,
        "start": start.isoformat(),
        "end": end.isoformat(),
        "cost_pips": cost_pips,
        "trades": len(trades),
        "wins": len(wins),
        "losses": len(losses),
        "win_rate": round(100.0 * len(wins) / len(trades), 2) if trades else None,
        "profit_factor": round(profit_factor, 3) if profit_factor is not None else None,
        "no_loss_pf": bool(trades and not losses),
        "net_return_pct": round((equity - 1.0) * 100.0, 2),
        "sum_return_pct": round(sum(returns), 2),
        "max_drawdown_pct": round(max_drawdown, 2),
        "avg_holding_days": round(statistics.mean(hold_values), 1) if hold_values else None,
    }
    return {"summary": summary, "equity": curve, "trades": trades}


def build_default_backtests(pairs: list[dict], rates: dict[str, dict[date, float]]) -> dict:
    results: list[dict] = []
    for pair in pairs:
        prices = derive_pair_prices(pair["pair"], rates)
        end = min(max(prices), datetime.now(ZoneInfo("America/Sao_Paulo")).date())
        result = backtest_pair(
            pair["pair"], pair["_fund_history"], prices,
            DEFAULT_BACKTEST_START, end, DEFAULT_COST_PIPS,
        )
        result["summary"]["validation"] = (
            "PROVISIONAL_PIT" if any(code in pair["pair"] for code in ("AUD", "CHF"))
            else "PIT_CAUSAL"
        )
        pair["backtest"] = result["summary"]
        results.append(result)
    results.sort(
        key=lambda result: (
            result["summary"]["profit_factor"] is not None,
            result["summary"]["profit_factor"] or -1.0,
            result["summary"]["net_return_pct"],
        ),
        reverse=True,
    )
    return {
        "meta": {
            "engine": "FUND_DIRECIONAL_D1",
            "start": DEFAULT_BACKTEST_START.isoformat(),
            "end": max(result["summary"]["end"] for result in results),
            "cost_pips": DEFAULT_COST_PIPS,
            "entry": "FUND leaves NEUTRAL; execution at the next ECB fixing",
            "exit": "FUND loses a band in the direction of the position",
            "scope_warning": "Nao inclui BO, REGIAO, ZOI, SL ATR ou dados intraday.",
        },
        "results": results,
    }


def run_backtest_request(payload: dict) -> dict:
    pair_name = str(payload.get("pair", "ALL")).upper()
    if pair_name != "ALL" and pair_name not in PAIR_ORDER:
        raise ValueError("par invalido")
    start = parse_iso(str(payload.get("start", DEFAULT_BACKTEST_START.isoformat())))
    end = parse_iso(str(payload.get("end", date.today().isoformat())))
    if start >= end:
        raise ValueError("data inicial deve ser anterior a data final")
    cost_pips = float(payload.get("cost_pips", DEFAULT_COST_PIPS))
    if not 0.0 <= cost_pips <= 25.0:
        raise ValueError("custo deve ficar entre 0 e 25 pips")

    if pair_name == "ALL" and start == DEFAULT_BACKTEST_START and cost_pips == DEFAULT_COST_PIPS and BACKTEST_PATH.exists():
        cached = json.loads(BACKTEST_PATH.read_text(encoding="utf-8"))
        if end == parse_iso(cached["meta"]["end"]):
            return cached

    series = read_all_yields()
    pit = {currency: shift_pit(values) for currency, values in series.items()}
    rates = read_fx_rates()
    names = PAIR_ORDER if pair_name == "ALL" else (pair_name,)
    results = []
    for name in names:
        pair = compute_pair(name, pit[name[:3]], pit[name[3:]])
        prices = derive_pair_prices(name, rates)
        result_end = min(end, max(prices))
        result = backtest_pair(name, pair["_fund_history"], prices, start, result_end, cost_pips)
        result["summary"]["validation"] = (
            "PROVISIONAL_PIT" if any(code in name for code in ("AUD", "CHF"))
            else "PIT_CAUSAL"
        )
        results.append(result)
    results.sort(
        key=lambda result: (
            result["summary"]["profit_factor"] is not None,
            result["summary"]["profit_factor"] or -1.0,
            result["summary"]["net_return_pct"],
        ),
        reverse=True,
    )
    return {
        "meta": {
            "engine": "FUND_DIRECIONAL_D1",
            "start": start.isoformat(),
            "end": end.isoformat(),
            "cost_pips": cost_pips,
            "entry": "FUND leaves NEUTRAL; execution at the next ECB fixing",
            "exit": "FUND loses a band in the direction of the position",
            "scope_warning": "Nao inclui BO, REGIAO, ZOI, SL ATR ou dados intraday.",
        },
        "results": results,
    }


def quality(currency: str, series: dict[date, float | None], today: date) -> dict:
    valid = {day: value for day, value in series.items() if value is not None}
    latest = max(valid)
    age = business_age(latest, today)
    definition = SOURCES[currency]
    if age <= definition["current_bdays"]:
        status = "CURRENT"
    elif age <= definition["delayed_bdays"]:
        status = "DELAYED"
    else:
        status = "STALE"
    return {
        "currency": currency,
        "status": status,
        "first_observation": min(valid).isoformat(),
        "last_observation": latest.isoformat(),
        "business_days_old": age,
        "yield_2y": round(valid[latest], 4),
        "source": definition["name"],
        "source_url": definition["url"],
        "cadence": definition["cadence"],
        "route": definition["route"],
        "observations": len(valid),
    }


def pair_quality(pair: dict, qualities: dict[str, dict]) -> None:
    base_state = qualities[pair["base"]]["status"]
    quote_state = qualities[pair["quote"]]["status"]
    if base_state == "CURRENT" and quote_state == "CURRENT":
        pair["data_status"] = "CURRENT"
        pair["operational"] = True
    elif "STALE" in (base_state, quote_state):
        pair["data_status"] = "STALE"
        pair["operational"] = False
    else:
        pair["data_status"] = "DELAYED"
        pair["operational"] = False

    if not pair["operational"]:
        pair["decision"] = "DADO_BLOQUEADO"
    elif pair["fund"] >= 25:
        pair["decision"] = "COMPRAR_BASE"
    elif pair["fund"] <= -25:
        pair["decision"] = "VENDER_BASE"
    else:
        pair["decision"] = "NEUTRAL"


def build_currency_ranking(pairs: list[dict], qualities: dict[str, dict]) -> list[dict]:
    ranking: list[dict] = []
    for currency in CURRENCIES:
        signed_scores: list[float] = []
        for pair in pairs:
            if not pair["operational"] or currency not in (pair["base"], pair["quote"]):
                continue
            signed_scores.append(pair["fund"] if currency == pair["base"] else -pair["fund"])
        score = statistics.mean(signed_scores) if signed_scores else None
        ranking.append({
            "currency": currency,
            "score": round(score, 2) if score is not None else None,
            "strength": classify(score),
            "valid_crosses": len(signed_scores),
            "data_status": qualities[currency]["status"],
            "yield_2y": qualities[currency]["yield_2y"],
        })
    ranking.sort(key=lambda row: float("-inf") if row["score"] is None else row["score"], reverse=True)
    return ranking


def carry_on_business_days(values: dict[date, float], days: list[date], limit: int = FFILL_LIMIT) -> dict[date, float | None]:
    output: dict[date, float | None] = {}
    prior_days = [day for day in values if day < days[0]] if days else []
    prior = max(prior_days) if prior_days else None
    last: float | None = values[prior] if prior is not None else None
    gap = max(0, business_age(prior, days[0]) - 1) if prior is not None else limit + 1
    for day in days:
        if day in values:
            last = values[day]
            gap = 0
        else:
            gap += 1
        output[day] = last if last is not None and gap <= limit else None
    return output


def attach_pre_fund_history(watchlist: dict[str, list[dict]]) -> int:
    """Anexa as observacoes PRE-FUND de cada dia aos calendarios ja gravados.

    Aditivo: nao altera nenhum campo existente. O campo novo por dia e
    'pre_fund_watch'. A selecao vem do walk-forward anual (causal); o bloco
    'outcome' de cada item e resultado observado depois, exibido apenas como
    conferencia - ele nunca participou da escolha.
    """
    if not watchlist:
        return 0
    total = 0
    for path in sorted(CALENDAR_DIR.glob("calendar_*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        touched = False
        for day in payload.get("days", []):
            items = watchlist.get(day.get("date"))
            if items:
                day["pre_fund_watch"] = items
                touched = True
                total += 1
            elif "pre_fund_watch" in day:
                del day["pre_fund_watch"]
                touched = True
        if touched:
            payload.setdefault("meta", {})["pre_fund_watch"] = (
                "observacoes PRE-FUND do proprio dia (walk-forward anual, causal); "
                "'outcome' e conferencia posterior e nao participou da selecao"
            )
            path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=1),
                encoding="utf-8",
            )
    return total


def build_daily_calendar(
    pairs: list[dict],
    pit: dict[str, dict[date, float | None]],
    rates: dict[str, dict[date, float]],
    today: date,
) -> dict:
    """Gera recomendacoes contemporaneas, sem PF futuro nem resultado posterior."""
    CALENDAR_DIR.mkdir(parents=True, exist_ok=True)
    weekdays = business_days(CALENDAR_START, today)
    available_yields: dict[str, dict[date, float | None]] = {}
    for currency in CURRENCIES:
        clean = {
            day: float(value)
            for day, value in pit[currency].items()
            if value is not None and day <= today
        }
        available_yields[currency] = carry_on_business_days(clean, weekdays)

    pair_daily: dict[str, dict[date, float | None]] = {}
    for pair in pairs:
        history = {
            parse_iso(row["date"]): float(row["fund"])
            for row in pair["_fund_history"]
            if parse_iso(row["date"]) <= today
        }
        pair_daily[pair["pair"]] = carry_on_business_days(history, weekdays)

    # Para datas historicas, a existencia do fixing ECB evita recomendar em
    # feriados reais (por exemplo, 1 de janeiro). O dia corrente continua
    # elegivel antes da publicacao do fixing, desde que seja segunda-sexta.
    historical_market_days = set(rates["USD"])

    by_year: dict[int, list[dict]] = {year: [] for year in range(CALENDAR_START.year, today.year + 1)}
    first_full: date | None = None
    full_days = partial_days = no_data_days = closed_days = 0
    cursor = CALENDAR_START
    while cursor <= today:
        if cursor.weekday() >= 5 or (cursor < today and cursor not in historical_market_days):
            by_year[cursor.year].append({
                "date": cursor.isoformat(),
                "market": "FECHADO",
                "coverage": "CLOSED",
                "available_currencies": 0,
                "valid_pairs": 0,
                "missing_currencies": [],
                "strongest": None,
                "weakest": None,
                "currencies": [],
                "recommendations": [],
            })
            closed_days += 1
            cursor += timedelta(days=1)
            continue

        available = [
            currency for currency in CURRENCIES
            if available_yields[currency].get(cursor) is not None
        ]
        missing = [currency for currency in CURRENCIES if currency not in available]
        day_pairs: list[dict] = []
        for pair in pairs:
            name = pair["pair"]
            score = pair_daily[name].get(cursor)
            if score is None or pair["base"] not in available or pair["quote"] not in available:
                continue
            day_pairs.append({
                "pair": name,
                "base": pair["base"],
                "quote": pair["quote"],
                "fund": round(score, 2),
                "strength": classify(score),
            })

        ranking: list[dict] = []
        for currency in CURRENCIES:
            signed_scores = [
                row["fund"] if row["base"] == currency else -row["fund"]
                for row in day_pairs
                if currency in (row["base"], row["quote"])
            ]
            score = statistics.mean(signed_scores) if signed_scores else None
            ranking.append({
                "currency": currency,
                "score": round(score, 2) if score is not None else None,
                "strength": classify(score),
                "valid_crosses": len(signed_scores),
                "yield_2y": round(available_yields[currency][cursor], 4)
                if available_yields[currency].get(cursor) is not None else None,
            })
        ranking.sort(
            key=lambda row: float("-inf") if row["score"] is None else row["score"],
            reverse=True,
        )
        ranked = [row for row in ranking if row["score"] is not None]
        positions = {row["currency"]: index + 1 for index, row in enumerate(ranked)}

        directional = [row for row in day_pairs if abs(row["fund"]) >= 25]
        directional.sort(key=lambda row: (abs(row["fund"]), row["pair"]), reverse=True)
        recommendations = []
        for index, row in enumerate(directional[:5]):
            long_side = row["fund"] > 0
            leader = row["base"] if long_side else row["quote"]
            laggard = row["quote"] if long_side else row["base"]
            recommendations.append({
                "rank": index + 1,
                "pair": row["pair"],
                "fund": row["fund"],
                "strength": row["strength"],
                "decision": "COMPRAR_BASE" if long_side else "VENDER_BASE",
                "leader": leader,
                "laggard": laggard,
                "reason": (
                    f"{leader} #{positions.get(leader, '-')} vs "
                    f"{laggard} #{positions.get(laggard, '-')} | FUND {row['fund']:+.1f}"
                ),
            })

        if len(available) == len(CURRENCIES) and len(day_pairs) == len(PAIR_ORDER):
            coverage = "STRICT_FULL"
            full_days += 1
            if first_full is None:
                first_full = cursor
        elif day_pairs:
            coverage = "PARTIAL"
            partial_days += 1
        else:
            coverage = "NO_DATA"
            no_data_days += 1

        by_year[cursor.year].append({
            "date": cursor.isoformat(),
            "market": "ABERTO",
            "coverage": coverage,
            "available_currencies": len(available),
            "valid_pairs": len(day_pairs),
            "missing_currencies": missing,
            "strongest": ranked[0] if ranked else None,
            "weakest": ranked[-1] if ranked else None,
            "currencies": ranking,
            "recommendations": recommendations,
        })
        cursor += timedelta(days=1)

    years = []
    for year, days in by_year.items():
        counts = {
            status: sum(day["coverage"] == status for day in days)
            for status in ("STRICT_FULL", "PARTIAL", "NO_DATA", "CLOSED")
        }
        payload = {
            "meta": {
                "year": year,
                "model": "FUND_V0.1_CAUSAL",
                "lookahead": False,
                "recommendation_rule": "abs(FUND) >= 25; ranked using only information available on that day",
                "counts": counts,
            },
            "days": days,
        }
        path = CALENDAR_DIR / f"calendar_{year}.json"
        # PRESERVA o que outros escritores anexaram por dia. Este bloco reconstroi
        # "days" do zero; sem o merge abaixo ele apagava "projection" e
        # "pair_funds", escritos por projection_history.py — foi o que aconteceu
        # em 21/ago/2026, quando esta funcao rodou 7 minutos depois dele e a aba
        # "Where the FUND could move next session" ficou vazia em todo o historico.
        if path.exists():
            try:
                antigo = json.loads(path.read_text(encoding="utf-8"))
                guardado = {
                    d.get("date"): {k: d[k] for k in ("projection", "pair_funds") if k in d}
                    for d in antigo.get("days", [])
                }
                for d in payload["days"]:
                    extra = guardado.get(d.get("date"))
                    if extra:
                        d.update(extra)
            except Exception:
                pass          # arquivo corrompido ou inexistente: segue sem merge
        path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        years.append({"year": year, "file": f"calendar_{year}.json", "counts": counts})

    index = {
        "meta": {
            "start": CALENDAR_START.isoformat(),
            "end": today.isoformat(),
            "first_strict_full": first_full.isoformat() if first_full else None,
            "currencies": list(CURRENCIES),
            "pairs": len(PAIR_ORDER),
            "lookahead": False,
            "partial_policy": "Shows only pairs whose both legs and FUND existed on that day.",
            "full_policy": "STRICT_FULL requires 8 currencies and 28 pairs with a causal reading.",
            "weekends": "CLOSED; no recommendation.",
            "full_days": full_days,
            "partial_days": partial_days,
            "no_data_days": no_data_days,
            "closed_days": closed_days,
        },
        "years": years,
    }
    CALENDAR_INDEX_PATH.write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return index


def build_snapshot(downloads: list[DownloadResult]) -> dict:
    series = read_all_yields()
    today = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
    qualities = {currency: quality(currency, values, today) for currency, values in series.items()}
    pit = {currency: shift_pit(values) for currency, values in series.items()}

    pairs: list[dict] = []
    for pair_name in PAIR_ORDER:
        pair = compute_pair(pair_name, pit[pair_name[:3]], pit[pair_name[3:]])
        pair_quality(pair, qualities)
        pairs.append(pair)

    rates = read_fx_rates()
    default_backtests = build_default_backtests(pairs, rates)
    calendar_index = build_daily_calendar(pairs, pit, rates, today)
    prices_by_pair = {
        pair["pair"]: derive_pair_prices(pair["pair"], rates)
        for pair in pairs
    }
    next_day = build_next_day_observations(
        pairs,
        prices_by_pair,
        today,
    )
    pre_fund = build_pre_fund_radar(pairs, prices_by_pair, today)
    dias_com_watch = attach_pre_fund_history(pre_fund.pop("history", {}))
    print(f"PRE-FUND historico anexado a {dias_com_watch} dias do calendario")
    priority_pairs = [
        pair for pair in pairs
        if pair["operational"]
        and abs(pair["fund"]) >= 25
        and pair["backtest"]["trades"] >= 10
        and pair["backtest"]["validation"] == "PIT_CAUSAL"
        and (pair["backtest"]["profit_factor"] or 0.0) > 1.0
    ]
    priority_pairs.sort(
        key=lambda pair: (
            pair["backtest"]["profit_factor"] is not None,
            pair["backtest"]["profit_factor"] or -1.0,
            abs(pair["fund"]),
        ),
        reverse=True,
    )
    priorities = [
        {
            "rank": index + 1,
            "pair": pair["pair"],
            "fund": pair["fund"],
            "decision": pair["decision"],
            "profit_factor": pair["backtest"]["profit_factor"],
            "trades": pair["backtest"]["trades"],
            "win_rate": pair["backtest"]["win_rate"],
            "net_return_pct": pair["backtest"]["net_return_pct"],
            "max_drawdown_pct": pair["backtest"]["max_drawdown_pct"],
        }
        for index, pair in enumerate(priority_pairs)
    ]
    news = read_news(pairs)
    for pair in pairs:
        pair.pop("_fund_history", None)

    ranking = build_currency_ranking(pairs, qualities)
    valid_pairs = [pair for pair in pairs if pair["operational"]]
    aligned = [pair for pair in valid_pairs if pair["decision"] in ("COMPRAR_BASE", "VENDER_BASE")]
    now = datetime.now(ZoneInfo("America/Sao_Paulo"))
    return {
        "meta": {
            "title": "HCI FUND Radar",
            "generated_at": now.isoformat(timespec="seconds"),
            "timezone": "America/Sao_Paulo",
            "currencies": len(CURRENCIES),
            "pairs": len(PAIR_ORDER),
            "operational_pairs": len(valid_pairs),
            "aligned_pairs": len(aligned),
            "warning": "FUND picks the side; timing, BO, REGION and ZOI remain manual decisions.",
            "backtest_warning": "AUD uses the RBA series via mirror with weekly publication; this snapshot does not replace a historical PIT backtest dataset.",
            "news_count": len(news),
        },
        "methodology": {
            "version": "FUND V0.1 congelado",
            "lookback": LOOKBACK,
            "normalization_window": NORM_WINDOW,
            "minimum_history": MIN_HISTORY,
            "pit_lag": "+1 dia util",
            "formula": "100 * tanh(((momento20 - media252_ex_ante) / desvio252_ex_ante) / 2)",
            "thresholds": {"strong": 60, "directional": 25},
            "exit_rule": "Exit when FUND loses a strength band in the direction of the position.",
        },
        "downloads": [result.__dict__ for result in downloads],
        "sources": [qualities[currency] for currency in CURRENCIES],
        "currencies": ranking,
        "priorities": priorities,
        "pairs": pairs,
        "news": news,
        "news_source": {
            "name": "Forex Factory / Fair Economy Media",
            "url": "https://www.forexfactory.com/calendar",
            "scope": "semana corrente; confirmar horario e release na fonte oficial",
        },
        "calendar": calendar_index["meta"],
        "backtest": default_backtests["meta"],
        "next_day": next_day,
        "pre_fund": pre_fund,
        "_backtest_default": default_backtests,
    }


def main(argv: list[str] | None = None) -> dict:
    parser = argparse.ArgumentParser(description="Atualiza o HCI FUND Radar")
    parser.add_argument("--force", action="store_true", help="ignora TTL do cache")
    parser.add_argument("--offline", action="store_true", help="nao acessa a internet")
    args = parser.parse_args(argv)

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    downloads = refresh_downloads(force=args.force, offline=args.offline)
    for result in downloads:
        log(f"{result.currency}: {result.state} | {result.detail}")
    failed = [result.currency for result in downloads if result.state == "FALHOU"]
    if failed:
        raise RuntimeError("fontes sem cache: " + ", ".join(failed))

    snapshot = build_snapshot(downloads)
    backtest_default = snapshot.pop("_backtest_default")
    BACKTEST_PATH.write_text(
        json.dumps(backtest_default, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    SNAPSHOT_PATH.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    log(
        f"snapshot salvo: {SNAPSHOT_PATH} | "
        f"{snapshot['meta']['operational_pairs']}/{snapshot['meta']['pairs']} pares operacionais | "
        f"{snapshot['meta']['aligned_pairs']} com lado definido | "
        f"backtest {len(backtest_default['results'])} pares"
    )
    return snapshot


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"ERRO: {error}", file=sys.stderr)
        raise
